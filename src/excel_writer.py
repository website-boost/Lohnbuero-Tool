from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import FIELD_LABEL_ALIASES, EmployeeRecord


# First data row in the template (rows 1–15 are header band in the standard
# Insolvenzgeld template — row 16 is the first employee row).
FIRST_DATA_ROW = 16
SHEET_NAME = "Arbeitnehmer"

# How many rows above FIRST_DATA_ROW to scan for column headers.
# 8 covers all known templates (header band rows 8–15).
HEADER_SCAN_ROWS = 8

ORANGE_FILL = PatternFill(start_color="FFD9A0", end_color="FFD9A0", fill_type="solid")


@dataclass
class WriteResult:
    """Outcome of a write_records call — useful for the GUI to surface warnings."""
    rows_written: int = 0
    detected_columns: dict[str, str] = field(default_factory=dict)  # field → col letter
    fields_without_column: list[str] = field(default_factory=list)  # extracted but no col


def _normalize(s: str) -> str:
    """Loose normalization for header matching.

    Lowercases, strips parentheticals, collapses German umlauts to ae/oe/ue/ss,
    and removes ALL whitespace and punctuation. So 'Pers-Nr.', 'pers nr',
    'Personalnr', 'PERS_NR' all collapse to 'persnr'.
    """
    s = s.lower()
    s = re.sub(r"\(.*?\)", "", s)  # drop parentheticals like "(Nachname)"
    s = (s.replace("ä", "ae").replace("ö", "oe")
           .replace("ü", "ue").replace("ß", "ss"))
    s = re.sub(r"[\s\-./_,:#§*]+", "", s)
    return s


def detect_column_map(ws: Worksheet) -> dict[str, str]:
    """Scan the header band of the worksheet → return field_name → column letter.

    Combines all non-empty cells per column in rows (FIRST_DATA_ROW - HEADER_SCAN_ROWS)
    .. (FIRST_DATA_ROW - 1) into one normalized blob per column, then matches against
    each field's known aliases. Strict equality only — no fuzzy substring matches
    (which would conflate 'Name' with 'Nachname' subspans, or 'Kinder' with everything).
    """
    start_row = max(1, FIRST_DATA_ROW - HEADER_SCAN_ROWS)
    end_row = FIRST_DATA_ROW - 1

    # column letter → list of normalized text fragments seen in that column
    column_texts: dict[str, list[str]] = {}
    for row in range(start_row, end_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            normed = _normalize(text)
            if not normed:
                continue
            col_letter = get_column_letter(col_idx)
            column_texts.setdefault(col_letter, []).append(normed)

    # Pre-normalize all aliases.
    field_to_normed_aliases: dict[str, set[str]] = {
        field: {_normalize(alias) for alias in aliases}
        for field, aliases in FIELD_LABEL_ALIASES.items()
    }

    detected: dict[str, str] = {}
    used_columns: set[str] = set()

    # For each column, try each field's alias set. Match if any normalized fragment
    # in that column equals any normalized alias of the field.
    for col_letter, fragments in column_texts.items():
        if col_letter in used_columns:
            continue
        for field_name, aliases in field_to_normed_aliases.items():
            if field_name in detected:
                continue
            if any(frag in aliases for frag in fragments):
                detected[field_name] = col_letter
                used_columns.add(col_letter)
                break

    return detected


def write_records(
    template_path: Path,
    output_path: Path,
    records: list[EmployeeRecord],
) -> WriteResult:
    """Copy the template, then fill rows 16+ with employee data.

    Columns are detected from the template's header band — so the same code
    works whether the template has 18 columns (basic Stammdaten) or 30+
    columns (with Bankdaten, Steuerklasse, Krankenkasse, etc.).

    Cells where extraction was uncertain are filled orange and get a comment
    showing the conflicting values, so the user knows where to double-check.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Template missing expected sheet '{SHEET_NAME}'")
    ws = wb[SHEET_NAME]

    column_map = detect_column_map(ws)
    if not column_map:
        raise ValueError(
            f"Konnte keine bekannten Spaltenüberschriften im Blatt '{SHEET_NAME}' finden. "
            f"Erwartet werden Header wie 'Pers-Nr', 'Name', 'Vorname' usw. in den Zeilen "
            f"{FIRST_DATA_ROW - HEADER_SCAN_ROWS}–{FIRST_DATA_ROW - 1}."
        )

    # Surface which extracted fields couldn't be placed (no matching column).
    written_field_names: set[str] = set()

    for offset, record in enumerate(records):
        row = FIRST_DATA_ROW + offset
        for field_name, column_letter in column_map.items():
            value = record.data.get(field_name)
            if value is None or value == "":
                continue
            cell = ws[f"{column_letter}{row}"]
            cell.value = value
            written_field_names.add(field_name)
            if field_name in record.uncertain_fields:
                cell.fill = ORANGE_FILL
                note = record.notes.get(field_name, "Bitte prüfen — unsicher extrahiert.")
                cell.comment = Comment(note, "Lohnbuero-Tool")

    # Fields that records had data for but no column existed in the template.
    fields_without_column: list[str] = []
    for record in records:
        for field_name in FIELD_LABEL_ALIASES:
            if record.data.get(field_name) and field_name not in column_map:
                if field_name not in fields_without_column:
                    fields_without_column.append(field_name)

    wb.save(output_path)
    return WriteResult(
        rows_written=len(records),
        detected_columns=column_map,
        fields_without_column=fields_without_column,
    )
